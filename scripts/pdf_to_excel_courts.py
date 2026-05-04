"""법원구성부.pdf → 엑셀 변환.

PDF의 재판부 구성표를 단어 좌표 기반으로 파싱해서 (법원, 부별, 직위, 성명) 레코드
구조의 엑셀 파일로 저장한다. pdfplumber의 extract_tables() 가 한 셀 안에 여러 행을
newline으로 합쳐버리는 한계가 있어, x좌표로 컬럼을 정의하고 y좌표로 행을
재구성하는 방식을 쓴다. 1페이지(대법원)는 7열 레이아웃이라 별도 파서를 둔다.

출력 시트:
- 전체     : 모든 법원 통합 (법원, 부별, 직위, 성명)
- 대법원    : 대법원만 직위·부호별 정렬
- 법원별 시트들 : 법원마다 한 시트 (부별, 직위, 성명)
"""
import re
import sys
import pdfplumber
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PDF_PATH = r"C:\Users\hyun\Desktop\법원구성부.pdf"
OUTPUT_PATH = r"C:\Users\hyun\Desktop\법원구성부_정제.xlsx"

# 6열(부별/직위/성명 ×2) 컬럼 경계 (관찰값 기반).
# 헤더 x0: 부별1=103, 직위1=190, 성명1=245, 부별2=321, 직위2=408, 성명2=463.
# 셀 중심 cx 기준.
COL_BOUNDS = [
    ("L_부별", 0,    165),
    ("L_직위", 165,  230),
    ("L_성명", 230,  290),
    ("R_부별", 290,  380),
    ("R_직위", 380,  445),
    ("R_성명", 445,  10000),
]

COURT_KEYWORDS = (
    "대법원", "고등법원", "지방법원", "특허법원",
    "가정법원", "행정법원", "회생법원",
    "재판부", "지원",
)


def assign_col(cx):
    for i, (_n, lo, hi) in enumerate(COL_BOUNDS):
        if lo <= cx < hi:
            return i
    return None


def collapse_korean_spaces(s):
    """한글 글자 사이 공백을 제거. (법 원 장 -> 법원장, 판 사 -> 판사)
    한글 외 문자(괄호, 숫자, 영문)와 한글 사이 공백은 보존하지 않아도 무방하나
    안전하게 모든 공백 제거 후 정리.
    """
    if not s:
        return ""
    s = re.sub(r"\s+", "", s)
    return s.strip()


def extract_page_rows(page):
    """페이지에서 (cells_6) 행 리스트 반환. 각 행은 6개 문자열."""
    words = page.extract_words(x_tolerance=2, y_tolerance=2, keep_blank_chars=False)
    if not words:
        return []

    # 데이터 영역만 사용: 페이지 헤더(연도/페이지번호)와 푸터 페이지번호 제거
    # 관찰: 헤더 top<=70, 푸터 top>=페이지하단-50
    page_h = page.height
    data_words = [w for w in words if w["top"] > 100 and w["top"] < page_h - 60]

    # y로 클러스터링
    Y_TOL = 4
    rows = []  # list of (y_center, [words])
    for w in sorted(data_words, key=lambda w: w["top"]):
        placed = False
        for r in rows:
            if abs(r[0] - w["top"]) <= Y_TOL:
                r[1].append(w)
                # 평균 y 갱신
                ys = [x["top"] for x in r[1]]
                r[0] = sum(ys) / len(ys)
                placed = True
                break
        if not placed:
            rows.append([w["top"], [w]])

    rows.sort(key=lambda r: r[0])

    # 각 행을 6개 컬럼으로 분배
    structured = []
    for y, ws in rows:
        cells = ["", "", "", "", "", ""]
        # x0 순으로 정렬해야 한 셀 안에서 글자 순서 유지됨
        for w in sorted(ws, key=lambda w: w["x0"]):
            cx = (w["x0"] + w["x1"]) / 2
            ci = assign_col(cx)
            if ci is None:
                continue
            cells[ci] = (cells[ci] + " " + w["text"]).strip()
        # 각 셀 공백 정리
        cells = [collapse_korean_spaces(c) for c in cells]
        # 헤더 행(부별/직위/성명) 스킵
        if cells[0] == "부별" or cells[1] == "직위" or cells[2] == "성명":
            continue
        # 완전 공백 행 스킵
        if not any(cells):
            continue
        structured.append(cells)

    return structured


def split_left_right(rows):
    """좌측 묶음 → 우측 묶음 순서로 직렬화."""
    left = [(r[0], r[1], r[2]) for r in rows]
    right = [(r[3], r[4], r[5]) for r in rows]
    return left + right


_NAME_BLACKLIST = {
    "당직판사", "담당판사", "당직", "담당", "직무대리",
    "직무", "휴가", "공석", "미정",
    "대직제운영",  # 부별 라벨 노이즈
}


def clean_person_name(text):
    """성명 칸 정제: 노이즈/부적합한 이름 제거 + 정규화."""
    if not text:
        return ""
    s = str(text).strip()
    if all(c == "○" for c in s) and s:
        return ""
    if s.startswith(("(", "[")):
        return ""
    if not re.match(r"^[가-힣]", s):
        return ""
    if (")" in s or "]" in s) and "(" not in s and "[" not in s:
        return ""
    # '이종민(29)', '김창현(신청,' 등 → 한글 이름 부분만 추출
    m = re.match(r"^([가-힣]{2,5})(?:\s*[(\[]|$)", s)
    if m:
        s = m.group(1)
    # 직무명 (당직판사, 담당판사 등) 제외
    if s in _NAME_BLACKLIST:
        return ""
    return s


# 직위 정규화:
# (1) PDF 추출 시 잘림/합쳐짐 보정 + (2) 표준 직급 체계로 통합
# 표준 카테고리: 대법관, 법원장, 지원장, 부장판사, 고법판사, 판사, 원로법관,
#              수석재판연구관, 선임재판연구관, 재판연구관,
#              비서실장, 비서관, 참여사무관
_POSITION_NORMALIZE = {
    # (1) PDF 잘림/합쳐짐 보정
    "사": "판사",
    "부장판사판": "부장판사",
    "부장판사부장": "부장판사",
    "(부장)판사": "부장판사",

    # (2) 부장판사 변형 통합 (선임/수석/지법/민사·형사 수석 등 → 부장판사)
    "지법부장판사": "부장판사",
    "지법수석부장": "부장판사",
    "지법수석부장판사": "부장판사",
    "수석부장": "부장판사",
    "수석부장판사": "부장판사",
    "선임부장판사": "부장판사",
    "형사수석부장판사": "부장판사",
    "민사제1수석부장판사": "부장판사",
    "민사제2수석부장판사": "부장판사",
    "수석판사": "부장판사",

    # (3) 고법판사 변형 통합
    "고법수석판사": "고법판사",

    # (4) 재판연구관 변형 통합 (공동/파견·전문직 → 재판연구관)
    "공동재판연구관": "재판연구관",
    "파견·전문직재판연구관": "재판연구관",

    # (5) 비서관 변형 통합 (5급 표기 제거)
    "비서관(5급)": "비서관",
}


def normalize_position(pos):
    if not pos:
        return pos
    s = str(pos).strip()
    return _POSITION_NORMALIZE.get(s, s)


def is_continuation_label(bu_text):
    """라벨이 두 줄로 split된 뒤쪽 조각인지 판단.
    예: '법인파산)' (닫는 괄호만), '장애인·선거]' (닫는 대괄호만), '공보,사법행정지원)'."""
    if not bu_text:
        return False
    op = bu_text.count("(")
    cp = bu_text.count(")")
    ob = bu_text.count("[")
    cb = bu_text.count("]")
    return cp > op or cb > ob


def looks_like_court_header_loose(bu, jik, name):
    """좀 더 관대한 법원 헤더 판단: 부별만 있고 직위/성명 없음 + 키워드 포함."""
    if not bu:
        return False
    if jik or name:
        return False
    # 괄호/대괄호로 시작하는 텍스트는 부별 sub-label
    if bu.startswith(("(", "[")):
        return False
    # 줄바꿈된 라벨의 뒤쪽 조각 (닫는 괄호가 더 많은 경우) 도 법원 헤더 아님
    if is_continuation_label(bu):
        return False
    return any(kw in bu for kw in COURT_KEYWORDS)


_OUTPOST_RE = re.compile(r"^원외재판부\s*\(\s*([가-힣]+)\s*\)\s*$")


def detect_outpost_subcourt(bu):
    """'원외재판부(인천)' → '인천재판부'. 본원에 합쳐서 sub-court 컨텍스트로 사용."""
    if not bu:
        return None
    m = _OUTPOST_RE.match(bu)
    if m:
        return f"{m.group(1)}재판부"
    return None


def is_primary_court_name(name):
    """음영 헤더 텍스트가 본원(재판부 미포함) 인지."""
    if not name:
        return False
    return "재판부" not in name


def build_records(stream):
    """직렬 시퀀스 → (법원, 부별, 직위, 성명) 레코드.
    - 법원 헤더: 다음 데이터들의 법원 컨텍스트로 사용
    - top_court 추적: 본원(재판부 라벨 없는 법원). 원외재판부(X) 발견 시 'top_court X재판부'로 sub-court 합성
    - 부별 forward-fill
    - 괄호/대괄호 시작 보조 라벨은 현재 부별에 누적
    - 보조 라벨 등장 시 같은 부 이전 레코드들의 부별도 함께 갱신(backfill)
    """
    records = []
    current_court = None
    top_court = None  # 본원 (예: 서울고등법원). 원외재판부(X) 합성에 사용
    current_bu = None
    current_bu_indices = []  # 현재 부에 속한 records의 인덱스

    def maybe_apply_sublabel(bu_text):
        """bu_text가 보조 라벨/연속 라벨이면 current_bu에 누적하고 backfill. 변경 여부 반환."""
        nonlocal current_bu
        is_sublabel = (bu_text.startswith("(") or bu_text.startswith("[")
                       or is_continuation_label(bu_text))
        if is_sublabel and current_bu and bu_text not in current_bu:
            new_bu = current_bu + " " + bu_text
            for idx in current_bu_indices:
                records[idx]["부별"] = new_bu
            current_bu = new_bu
            return True
        return is_sublabel

    for bu, jik, name in stream:
        if not (bu or jik or name):
            continue

        # 0. 원외재판부(X) sub-court 전환 (본원 + X재판부로 합성)
        sub = detect_outpost_subcourt(bu)
        if sub:
            base = top_court or current_court or ""
            current_court = (base + " " + sub).strip()
            current_bu = None
            current_bu_indices = []
            # 이 행에 데이터가 있다면 sub-court 최상위 인원으로 기록
            if jik or name:
                records.append({
                    "법원": current_court,
                    "부별": "",
                    "직위": jik,
                    "성명": name,
                })
                current_bu_indices.append(len(records) - 1)
            continue

        # 1. 법원 헤더 후보: 부별만 있고 직위/성명 없음
        if bu and not jik and not name:
            if looks_like_court_header_loose(bu, jik, name):
                current_court = bu
                # 본원(primary)이면 top_court 갱신, 결합형(예: '대전고등법원 청주재판부')이면 유지
                if is_primary_court_name(bu):
                    top_court = bu
                current_bu = None
                current_bu_indices = []
                continue
            # 보조 라벨 또는 새 부별 단독 행
            if not maybe_apply_sublabel(bu):
                current_bu = bu
                current_bu_indices = []
            continue

        # 2. 직위 또는 성명이 있는 데이터 행
        if bu:
            if not maybe_apply_sublabel(bu):
                current_bu = bu
                current_bu_indices = []
        # 성명 정제: 노이즈 행 제거
        clean = clean_person_name(name)
        if not clean:
            continue
        records.append({
            "법원": current_court or "",
            "부별": current_bu or "",
            "직위": normalize_position(jik),
            "성명": clean,
        })
        current_bu_indices.append(len(records) - 1)

    # 직위 forward-fill: 같은 (법원, 부) 컨텍스트에서 직위가 비어있으면 직전 행의 직위 채택
    prev_jik = None
    prev_ctx = None
    for r in records:
        ctx = (r["법원"], r["부별"])
        if r["직위"]:
            prev_jik = r["직위"]
            prev_ctx = ctx
        else:
            if ctx == prev_ctx and prev_jik:
                r["직위"] = prev_jik

    return records


_SC_HEADER_TEXTS = {
    "직위", "주심부", "부호", "대법원장", "비서실", "수석재판연구관",
    "선임재판연구관", "재판연구관", "참여사무관", "공동재판연구관",
    "대법관", "민사", "형사", "특별", "비서관", "비서관(5급)", "비서실장",
    "파견", "전문직", "재판연구관" ,
}
_SC_BUHO = set("차아바자나마다사라가타카")  # 부호 한 글자
_SC_DIVISION_PARENS = re.compile(r"^\([가-힣·]+\)$")  # (차), (아·바·자), (차·자)


def _build_buho_to_bu_map(words):
    """부호 12개 (차/아/바/자/나/마/다/사/라/가/타/카)의 top 좌표를 정렬해 1부/2부/3부 매핑."""
    buho = []  # (top, 부호)
    for w in words:
        if 50 < w["x0"] < 105 and w["top"] > 250 and w["text"] in _SC_BUHO:
            buho.append((w["top"], w["text"]))
    buho.sort()
    mapping = {}  # round(top) → '1부' 등
    for i, (t, _bh) in enumerate(buho[:12]):
        mapping[round(t)] = "1부" if i < 4 else ("2부" if i < 8 else "3부")
    return mapping


def parse_supreme_court_records(page):
    """대법원 페이지를 단어 좌표로 파싱하여 (법원, 부별, 직위, 성명) records 반환.
    - 비서실 영역(top<250)
    - 대법관 표(top>=250): 부호 12개를 4개씩 묶어 1부/2부/3부 결정
    - 파견·전문직 재판연구관 헤더(top≈527) 이후 우측 컬럼 → 직위 분리
    """
    words = page.extract_words(x_tolerance=2, y_tolerance=2)
    page_h = page.height

    # 부호 → 부 매핑 사전 구축
    buho_to_bu = _build_buho_to_bu_map(words)

    # 파견·전문직 재판연구관 헤더 위치 (이후 우측 컬럼 인원은 별도 직위)
    pj_top = None
    for w in words:
        if "파견" in w["text"] and "전문직" in w["text"] and w["x0"] >= 340:
            pj_top = w["top"]
            break

    Y_TOL = 5
    raw_rows = []
    for w in sorted(words, key=lambda w: w["top"]):
        if w["top"] < 100 or w["top"] > page_h - 50:
            continue
        placed = False
        for r in raw_rows:
            if abs(r[0] - w["top"]) <= Y_TOL:
                r[1].append(w)
                ys = [x["top"] for x in r[1]]
                r[0] = sum(ys) / len(ys)
                placed = True
                break
        if not placed:
            raw_rows.append([w["top"], [w]])
    raw_rows.sort()

    def col_of(cx):
        if cx < 105: return "부"
        if cx < 150: return "대법관"
        if cx < 225: return "재판연구관"
        if cx < 265: return "비서관"
        if cx < 340: return "참여사무관"
        return "공동재판연구관"

    NOISE_TOKENS = {
        "아바자", "차자", "수석부", "장애인",
        "민민", "사사", "형형", "특특", "별별",
        "민사", "형사", "특별", "재판연구",
        "비서실", "주심부",
    }

    records = []

    for y, ws in raw_rows:
        cols = {"부": [], "대법관": [], "재판연구관": [], "비서관": [],
                "참여사무관": [], "공동재판연구관": []}
        for w in sorted(ws, key=lambda w: w["x0"]):
            cx = (w["x0"] + w["x1"]) / 2
            cols[col_of(cx)].append(w["text"])

        # ===== 상단 비서실 영역 =====
        if y < 250:
            row_text = collapse_korean_spaces(" ".join(
                cols["대법관"] + cols["재판연구관"] + cols["비서관"] +
                cols["참여사무관"] + cols["공동재판연구관"]))
            jik = None
            for t in ("비서실장", "비서관(5급)", "수석재판연구관", "선임재판연구관"):
                if t in row_text:
                    jik = t
                    break
            if jik:
                idx = row_text.find(jik)
                rest = row_text[idx + len(jik):]
                m = re.search(r"[가-힣]{2,4}", rest)
                if m:
                    name = m.group(0)
                    if name not in _SC_HEADER_TEXTS:
                        bu = "비서실" if "비서실" in jik or jik == "비서실장" else ""
                        records.append({
                            "법원": "대법원", "부별": bu,
                            "직위": normalize_position(jik), "성명": name,
                        })
            continue

        # ===== 하단 대법관 표 영역 =====
        # 부호 추출
        부호 = ""
        부호_top = None
        for t in cols["부"]:
            if t in _SC_BUHO:
                부호 = t
                break
        # 부호의 top 좌표는 행 y와 거의 동일
        if 부호:
            # 부호의 정확한 top 찾기
            for w in ws:
                if w["text"] == 부호 and 50 < w["x0"] < 105:
                    부호_top = w["top"]
                    break
        # 부 결정: 부호의 top으로 buho_to_bu 사전 조회
        bu_num = ""
        if 부호_top is not None:
            # round(부호_top) 기준 ±2 매칭
            for k, v in buho_to_bu.items():
                if abs(k - 부호_top) <= 2:
                    bu_num = v
                    break
        bu_label = (bu_num + (" " + 부호 if 부호 else "")).strip()

        # 대법관
        for w_text in cols["대법관"]:
            tok = collapse_korean_spaces(w_text)
            if tok in _SC_HEADER_TEXTS or _SC_DIVISION_PARENS.match(tok) or tok in NOISE_TOKENS:
                continue
            if re.match(r"^[가-힣]{2,4}$", tok):
                if clean_person_name(tok):
                    records.append({
                        "법원": "대법원", "부별": bu_label,
                        "직위": "대법관", "성명": tok,
                    })

        # 재판연구관/비서관/참여사무관 (좌측 컬럼들 - 부호별 매핑)
        for col_name in ("재판연구관", "비서관", "참여사무관"):
            cell = " ".join(cols[col_name])
            if not cell or collapse_korean_spaces(cell) in _SC_HEADER_TEXTS:
                continue
            for m in re.finditer(r"[가-힣]{2,4}", cell):
                tok = m.group(0)
                if tok in _SC_HEADER_TEXTS or tok in NOISE_TOKENS:
                    continue
                if len(tok) == 2 and tok[0] == tok[1]:
                    continue
                records.append({
                    "법원": "대법원", "부별": bu_label,
                    "직위": normalize_position(col_name), "성명": tok,
                })

        # 공동재판연구관 (부 attribution 없음, 파견·전문직 영역 분리)
        cell = " ".join(cols["공동재판연구관"])
        if cell and collapse_korean_spaces(cell) not in _SC_HEADER_TEXTS:
            # 파견·전문직 헤더 이후면 직위 변경
            jik = "파견·전문직재판연구관" if (pj_top and y > pj_top) else "공동재판연구관"
            for m in re.finditer(r"[가-힣]{2,4}", cell):
                tok = m.group(0)
                if tok in _SC_HEADER_TEXTS or tok in NOISE_TOKENS:
                    continue
                if len(tok) == 2 and tok[0] == tok[1]:
                    continue
                records.append({
                    "법원": "대법원", "부별": "",
                    "직위": normalize_position(jik), "성명": tok,
                })

    return records


def base_bu(bu):
    """합의부 주심 순환(제N-M부) 통합 → 제N부.
    단독부('단독' 포함)는 통합 대상 아님."""
    if not bu:
        return bu
    if "단독" in bu:
        return bu
    # "제N-M" → "제N" (M은 1자리 이상 숫자). 내부의 "(제N-M가사부)" 같은 표기도 함께 처리.
    return re.sub(r"(제\d+)-\d+", r"\1", bu)


def dedup_records(records):
    """중복 정제: (법원, base_부, 직위, 성명) 기준 unique. 부별은 base_bu 적용."""
    seen = set()
    out = []
    for r in records:
        new_bu = base_bu(r["부별"])
        key = (r["법원"], new_bu, r["직위"], r["성명"])
        if key in seen:
            continue
        seen.add(key)
        out.append({"법원": r["법원"], "부별": new_bu, "직위": r["직위"], "성명": r["성명"]})
    return out


def main():
    print(f"Reading {PDF_PATH} ...", file=sys.stderr)

    with pdfplumber.open(PDF_PATH) as pdf:
        # 1페이지: 대법원 (7열 구조 — 별도 파서)
        supreme_records = parse_supreme_court_records(pdf.pages[0])

        # 2~끝: 일반 6열 파서 (부별/직위/성명 ×2)
        all_stream = []
        for i, page in enumerate(pdf.pages[1:], start=2):
            page_rows = extract_page_rows(page)
            stream = split_left_right(page_rows)
            all_stream.extend(stream)
            if i % 20 == 0:
                print(f"  parsed page {i}", file=sys.stderr)

        court_records_raw = build_records(all_stream)

    # 대법원 records를 앞에 합쳐서 합의부 주심 순환 통합 + 중복 제거
    court_records = dedup_records(supreme_records + court_records_raw)

    print(f"  대법원 records:        {len(supreme_records)}", file=sys.stderr)
    print(f"  각급 법원 records(raw): {len(court_records_raw)}", file=sys.stderr)
    print(f"  전체 records (정제후):  {len(court_records)}", file=sys.stderr)

    # 법원 분포 확인
    by_court = defaultdict(list)
    for rec in court_records:
        by_court[rec["법원"] or "(미분류)"].append(rec)
    print(f"  법원 수: {len(by_court)}", file=sys.stderr)
    for ct, rs in list(by_court.items())[:8]:
        print(f"    - {ct}: {len(rs)}건", file=sys.stderr)

    # ===== 엑셀 작성 =====
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def style_header(ws, n_cols):
        for col in range(1, n_cols + 1):
            c = ws.cell(row=1, column=col)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = border
        ws.freeze_panes = "A2"

    # Sheet 1: 전체
    ws_all = wb.active
    ws_all.title = "전체"
    headers_all = ["법원", "부별", "직위", "성명"]
    ws_all.append(headers_all)
    style_header(ws_all, len(headers_all))
    for rec in court_records:
        ws_all.append([rec["법원"], rec["부별"], rec["직위"], rec["성명"]])
    for i, w in enumerate([28, 38, 16, 14], start=1):
        ws_all.column_dimensions[get_column_letter(i)].width = w
    for row in ws_all.iter_rows(min_row=2, max_row=ws_all.max_row, max_col=4):
        for c in row:
            c.alignment = left_align if c.column in (1, 2) else center
            c.border = border

    # Sheet 2: 대법원 (정제된 records 기반, 직위별 정렬)
    ws_sc = wb.create_sheet("대법원")
    ws_sc.append(["부별", "직위", "성명"])
    style_header(ws_sc, 3)
    sc_recs = [r for r in court_records if r["법원"] == "대법원"]
    # 직위 순서: 비서실 → 수석/선임 → 대법관 → 재판연구관 → 비서관 → 참여사무관 → 공동재판연구관 → 파견·전문직
    jik_order = {
        "비서실장": 0, "비서관(5급)": 1,
        "수석재판연구관": 2, "선임재판연구관": 3,
        "대법관": 4, "재판연구관": 5, "비서관": 6, "참여사무관": 7,
        "공동재판연구관": 8, "파견·전문직재판연구관": 9,
    }
    def sc_sort_key(r):
        jo = jik_order.get(r["직위"], 99)
        bu = r["부별"] or ""
        # 부별: '1부 차' 처럼 시작하면 부 번호 + 부호 순
        m = re.match(r"^(\d+)부\s*(.?)", bu)
        if m:
            num = int(m.group(1))
            buho = m.group(2)
            buho_idx = "차아바자나마다사라가타카".find(buho) if buho else 99
            return (jo, num, buho_idx, r["성명"])
        return (jo, 99, 99, r["성명"])
    for r in sorted(sc_recs, key=sc_sort_key):
        ws_sc.append([r["부별"], r["직위"], r["성명"]])
    ws_sc.column_dimensions["A"].width = 14
    ws_sc.column_dimensions["B"].width = 22
    ws_sc.column_dimensions["C"].width = 14
    for row in ws_sc.iter_rows(min_row=2, max_row=ws_sc.max_row, max_col=3):
        for c in row:
            c.alignment = center
            c.border = border

    # Sheet 3+: 법원별
    used = {"전체", "대법원"}

    def safe_title(name):
        t = re.sub(r"[\\/*?:\[\]]", "_", name)[:31] or "기타"
        base = t
        i = 2
        while t in used:
            suf = f"_{i}"
            t = (base[: 31 - len(suf)] + suf)
            i += 1
        used.add(t)
        return t

    for court, recs in by_court.items():
        title = safe_title(court)
        ws = wb.create_sheet(title)
        ws.append(["부별", "직위", "성명"])
        style_header(ws, 3)
        for r in recs:
            ws.append([r["부별"], r["직위"], r["성명"]])
        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 14
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=3):
            for c in row:
                c.alignment = left_align if c.column == 1 else center
                c.border = border

    wb.save(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH}", file=sys.stderr)
    print(f"Total sheets: {len(wb.sheetnames)}", file=sys.stderr)


if __name__ == "__main__":
    main()
