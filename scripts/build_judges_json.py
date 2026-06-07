"""법원구성부_정제.xlsx → data/judges.json 재생성.

- 사람마다 PDF에서 첫 등장한 법원·부·직위만 채택 (대표 부 1개)
- 기존 judges.json에서 photoUrl/createdAt는 이름 매칭으로 보존
- courts.json과 매핑하여 courtId/courtRegion 채움
- 정렬: 법원 → 직위(표준 직급 순) → 성명
"""
import json
import re
import sys
from datetime import datetime, timezone
from openpyxl import load_workbook

XLSX = r"C:\Users\hyun\Desktop\법원구성부_정제.xlsx"
COURTS_JSON = r"C:\Users\hyun\Desktop\pansawatch\data\courts.json"
JUDGES_JSON = r"C:\Users\hyun\Desktop\pansawatch\data\judges.json"
OUT = JUDGES_JSON  # in-place 갱신

sys.stdout.reconfigure(encoding="utf-8")


def norm_name(s):
    return re.sub(r"\s+", "", str(s)).strip() if s else ""


def main():
    # 1. 법원 매핑 로드
    with open(COURTS_JSON, encoding="utf-8") as f:
        courts = json.load(f)
    court_by_norm = {norm_name(c["name"]): c for c in courts}

    # 2. 기존 judges.json (photoUrl 등 보존용)
    with open(JUDGES_JSON, encoding="utf-8") as f:
        old = json.load(f)
    old_by_name = {}
    for j in old:
        # 동명이인은 첫 항목 우선
        old_by_name.setdefault(j["name"], j)
    print(f"기존 judges.json: {len(old)}건 / 고유 이름 {len(old_by_name)}", file=sys.stderr)

    # 3. PDF Excel 로드
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["전체"]

    # (인명, 법원) 단위 첫 등장 (court, bu, jik) 채택
    # — 같은 이름 다른 법원 = 동명이인으로 분리
    # — 같은 법원 동명이인은 PDF disambiguator(예: '김창현(73년생)')가 이름에 포함되어 자동 분리
    seen = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        court, bu, jik, name = r
        if not name:
            continue
        name = str(name).strip()
        court_n = (court or "").strip()
        key = (name, court_n)
        if key in seen:
            continue
        seen[key] = {
            "court": court_n,
            "division": (bu or "").strip(),
            "position": (jik or "").strip(),
        }
    print(f"PDF 추출: (이름,법원) 단위 {len(seen)}", file=sys.stderr)

    # 4. 새 judges 리스트
    now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    new_judges = []
    matched_to_court = 0
    preserved_photo = 0
    for i, ((name, _court_key), info) in enumerate(seen.items(), start=1):
        court_name = info["court"]
        court = court_by_norm.get(norm_name(court_name))
        if court:
            matched_to_court += 1

        old_rec = old_by_name.get(name, {})
        if old_rec.get("photoUrl"):
            preserved_photo += 1

        rec = {
            "id": f"judge-{i}",
            "name": name,
            "courtId": court["id"] if court else (old_rec.get("courtId")),
            "court": court["name"] if court else court_name,
            "courtRegion": court["region"] if court else old_rec.get("courtRegion"),
            "position": info["position"],
            "division": info["division"],
            "photoUrl": old_rec.get("photoUrl"),
            "createdAt": old_rec.get("createdAt", now),
            "updatedAt": now,
        }
        new_judges.append(rec)

    # 5. 저장 (안정 정렬: 법원 → 직위 → 성명)
    POS_ORDER = {
        "대법관": 0, "법원장": 1, "수석부장판사": 2, "민사제1수석부장판사": 3,
        "민사제2수석부장판사": 4, "지법수석부장판사": 5, "수석판사": 6,
        "고법수석판사": 7, "지법부장판사": 8, "부장판사": 9, "고법판사": 10,
        "판사": 11, "수석재판연구관": 12, "선임재판연구관": 13, "재판연구관": 14,
        "공동재판연구관": 15, "파견·전문직재판연구관": 16,
        "비서실장": 90, "비서관(5급)": 91, "비서관": 92, "참여사무관": 93,
    }
    new_judges.sort(key=lambda r: (
        r["court"] or "",
        POS_ORDER.get(r["position"], 99),
        r["name"],
    ))
    # id 재부여 (정렬 후)
    for i, r in enumerate(new_judges, start=1):
        r["id"] = f"judge-{i}"

    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(new_judges, f, ensure_ascii=False, indent=2)

    print(f"\n저장 완료: {OUT}", file=sys.stderr)
    print(f"  총 {len(new_judges)}명", file=sys.stderr)
    print(f"  courtId 매핑 성공: {matched_to_court}/{len(new_judges)} ({matched_to_court/len(new_judges)*100:.1f}%)", file=sys.stderr)
    print(f"  photoUrl 보존: {preserved_photo}", file=sys.stderr)


if __name__ == "__main__":
    main()
